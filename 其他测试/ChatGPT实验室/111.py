import pandas as pd
from openpyxl import load_workbook
import os

# 定义源文件路径和工作表名
source_file_path = r"E:\SynologyDrive\Drive\24年\桌面-临时\跟线辅导表--自用\4月用\客户经销商信息20240412100623680516.xlsx"
sheet_name = 'sheet1'

# 读取源Excel文件的指定工作表
df = pd.read_excel(source_file_path, sheet_name=sheet_name)

# 定义需要提取的列，并按照新的顺序排列
columns_to_extract = [
    '客户名称', '订单额', '冰箱数量', '客户编码', '渠道类型',
    '上级经销商编码', '安排线路', '大区', '营业所', 'TDS', 'CR', '小区号'
]

# 从DataFrame中选择需要的列
data_to_insert = df[columns_to_extract]

# 定义模板文件路径
template_path = r"E:\SynologyDrive\Drive\24年\桌面-临时\跟线辅导表--自用\4月用\新的.xlsx"

# 检查文件是否存在，不存在则先创建一个空白的
if not os.path.exists(template_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Template"
    wb.save(template_path)

# 加载模板文件
book = load_workbook(template_path)

# 使用ExcelWriter时指定已
