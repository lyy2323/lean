import pandas as pd
from openpyxl import load_workbook

# 定义源文件路径和工作表名
source_file_path = r"E:\SynologyDrive\Drive\24年\桌面-临时\跟线辅导表--自用\4月用\客户经销商信息20240412100623680516.xlsx"
sheet_name = 'sheet1'

# 读取源Excel文件的指定工作表
df = pd.read_excel(source_file_path, sheet_name=sheet_name)

# 定义需要提取的列，并按照新的顺序排列
columns_to_extract = [
    '客户名称', '订单额', '冰箱数量', '客户编码', '渠道类型',
    '上级经销商编码', '安排线路', '大区', '营业所', 'TDS', 'CR', '小区号'
]

# 从DataFrame中选择需要的列
data_to_insert = df[columns_to_extract]

# 定义模板文件路径
template_path = r"E:\SynologyDrive\Drive\24年\桌面-临时\跟线辅导表--自用\4月用\新的.xlsx"

# 使用ExcelWriter时指定已加载的book对象，并且使用追加模式
with pd.ExcelWriter(template_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
    # 选择模板中的工作表，假设名为“Template”
    template_sheet_name = "Template"
    if template_sheet_name in writer.sheets:
        ws = writer.sheets[template_sheet_name]
        # 确保从第二行开始填充数据，避免覆盖头部信息
        max_row = ws.max_row
        for row in range(2, max_row + 1):
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).value = None  # 清除已有数据

    # 将数据写入模板工作表
    data_to_insert.to_excel(writer, sheet_name=template_sheet_name, startrow=1, index=False, header=False)

print("数据已成功填充至模板文件。")
