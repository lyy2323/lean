import pandas as pd
from openpyxl import load_workbook

# 设置数据文件路径
a_table_path = r"E:\SynologyDrive\Drive\24年\桌面-临时\跟线辅导表--自用\4月用\客户经销商信息20240412100623680516.xlsx"
pc_table_path = r"E:\SynologyDrive\Drive\24年\桌面-临时\跟线辅导表--自用\4月用\4月PC.xlsx"
output_path = r"E:\SynologyDrive\Drive\24年\桌面-临时\跟线辅导表--自用\4月用\新的.xlsx"

try:
    # 读取"A表"数据
    a_table = pd.read_excel(a_table_path, sheet_name='sheet1')

    # 定义需要提取的列
    columns_to_extract_a = ['客户名称', '订单额', '冰箱数量', '客户编码', '渠道类型',
                            '上级经销商编码', '安排线路', '大区', '营业所', 'TDS',
                            'CR', '小区号']
    data_to_insert_a = a_table[columns_to_extract_a]

    # 读取"PC清单"数据
    pc_table = pd.read_excel(pc_table_path, sheet_name='PC清单')

    # 定义需要从"PC清单"合并的列
    columns_to_merge = ['费用合计', '项目名称', '主货架', '主货架计划', '冰冻',
                        '冰冻计划', '第二陈列', '第二陈列计划', '其他', '其他计划',
                        '新增冰柜', '新增冰柜计划', '佳得乐', '佳得乐计划']

    # 合并数据
    merged_data = pd.merge(data_to_insert_a, pc_table[columns_to_merge + ['客户编码']],
                           on='客户编码', how='left')

    # 加载输出文件
    book = load_workbook(output_path)
    sheet = book['Template']  # 确保工作表名正确

    # 从第二行开始填充数据，保留格式
    for idx, row in merged_data.iterrows():
        for col_idx, value in enumerate(row, start=1):
            sheet.cell(row=idx + 2, column=col_idx, value=value)

    # 保存工作簿
    book.save(output_path)
    print("数据已成功填充至模板文件。")

except Exception as e:
    print("发生错误：", e)
